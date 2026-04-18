"""
PDDikti Dosen Explorer — Excel Exporter Service
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLOR_HEADER = "1F4E79"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ALT_ROW = "D6E4F0"
COLOR_TITLE = "2E75B6"


def _style_header(ws, row, max_col):
    fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    font = Font(name="Calibri", bold=True, color=COLOR_HEADER_FONT, size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin")
    )
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border


def _style_rows(ws, start, end, max_col):
    alt = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type="solid")
    border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin")
    )
    for r in range(start, end + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if (r - start) % 2 == 1:
                cell.fill = alt


def _auto_width(ws, max_col):
    for c in range(1, max_col + 1):
        mx = 0
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=True):
            for v in row:
                if v:
                    mx = max(mx, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(mx + 4, 55)


def export_dosen_excel(dosen_list, pt_map: dict) -> io.BytesIO:
    """Export dosen list to Excel and return as BytesIO buffer."""
    now_str = datetime.now().strftime("%d %B %Y, %H:%M WIB")
    total = len(dosen_list)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Dosen"

    cols = [
        "No", "Nama", "NIDN", "Perguruan Tinggi", "Jabatan Fungsional",
        "Status Ikatan Kerja", "Jenis Kelamin", "Program Studi",
        "Pendidikan Terakhir", "Status Aktivitas"
    ]
    MC = len(cols)

    # Title
    ws.merge_cells(f"A1:{get_column_letter(MC)}1")
    t = ws.cell(1, 1, "DATA DOSEN PROGRAM STUDI EKONOMI SYARIAH")
    t.font = Font(name="Calibri", bold=True, color=COLOR_TITLE, size=14)
    t.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"A2:{get_column_letter(MC)}2")
    s = ws.cell(2, 1, f"Sumber: PDDikti API — {now_str}")
    s.font = Font(italic=True, color="666666", size=9)
    s.alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A3:{get_column_letter(MC)}3")
    ws.cell(3, 1, f"Total Dosen: {total:,} | Program Studi Ekonomi Syariah").font = Font(bold=True, size=10)
    ws.cell(3, 1).alignment = Alignment(horizontal="center")

    # Header
    HR = 5
    for c, h in enumerate(cols, 1):
        ws.cell(HR, c, h)
    _style_header(ws, HR, MC)

    # Data rows
    for i, d in enumerate(dosen_list, 1):
        r = HR + i
        ws.cell(r, 1, i)
        ws.cell(r, 2, d.nama or "")
        ws.cell(r, 3, d.nidn or "")
        ws.cell(r, 4, pt_map.get(d.pt_id, ""))
        ws.cell(r, 5, d.jabatan_fungsional or "")
        ws.cell(r, 6, d.status_ikatan_kerja or "")
        ws.cell(r, 7, d.jenis_kelamin or "")
        ws.cell(r, 8, d.rumpun_prodi or "")
        ws.cell(r, 9, d.pendidikan_terakhir or "")
        ws.cell(r, 10, d.status_aktivitas or "")

    _style_rows(ws, HR + 1, HR + total, MC)
    _auto_width(ws, MC)
    ws.freeze_panes = "B6"
    ws.row_dimensions[1].height = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_prodi_detail_excel(prodi_list, pt_map: dict) -> io.BytesIO:
    """Export prodi detail list to Excel and return as BytesIO buffer."""
    now_str = datetime.now().strftime("%d %B %Y, %H:%M WIB")
    total = len(prodi_list)

    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Prodi"

    cols = [
        "No", "Nama Prodi", "Jenjang", "Perguruan Tinggi", "Jumlah Dosen",
        "Keterangan", "Akreditasi Program Studi", "PTN/PTS",
        "PTKIN/NON PTKIN", "DIKTI/DIKTIS", "Provinsi",
        "Semester Laporan Terakhir",
    ]
    MC = len(cols)

    # Title
    ws.merge_cells(f"A1:{get_column_letter(MC)}1")
    t = ws.cell(1, 1, "DAFTAR PROGRAM STUDI EKONOMI SYARIAH")
    t.font = Font(name="Calibri", bold=True, color=COLOR_TITLE, size=14)
    t.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"A2:{get_column_letter(MC)}2")
    s = ws.cell(2, 1, f"Sumber: PDDikti API — {now_str}")
    s.font = Font(italic=True, color="666666", size=9)
    s.alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A3:{get_column_letter(MC)}3")
    ws.cell(3, 1, f"Total Program Studi: {total:,}").font = Font(bold=True, size=10)
    ws.cell(3, 1).alignment = Alignment(horizontal="center")

    # Header
    HR = 5
    for c, h in enumerate(cols, 1):
        ws.cell(HR, c, h)
    _style_header(ws, HR, MC)

    # Data rows
    for i, p in enumerate(prodi_list, 1):
        r = HR + i
        ws.cell(r, 1, i)
        ws.cell(r, 2, p.nama_prodi or "")
        ws.cell(r, 3, p.jenjang or "")
        ws.cell(r, 4, pt_map.get(p.pt_id, ""))
        ws.cell(r, 5, p.jumlah_dosen or 0)
        ws.cell(r, 6, p.keterangan or "")
        ws.cell(r, 7, p.akreditasi or "")
        ws.cell(r, 8, p.ptn_pts or "")
        ws.cell(r, 9, p.ptkin_non_ptkin or "")
        ws.cell(r, 10, p.dikti_diktis or "")
        ws.cell(r, 11, p.provinsi or "")
        ws.cell(r, 12, p.semester_terakhir or "")

    _style_rows(ws, HR + 1, HR + total, MC)
    _auto_width(ws, MC)
    ws.freeze_panes = "B6"
    ws.row_dimensions[1].height = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
